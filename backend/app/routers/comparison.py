from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from typing import Optional
from io import BytesIO
from fastapi.responses import StreamingResponse

from ..services.storage import get_storage, StorageService
from ..services.openai_analyzer import OpenAIAnalyzer

router = APIRouter(tags=["comparison", "export"])


class ComparisonRequest(BaseModel):
    document_id_1: str
    document_id_2: str


class ComparisonDifference(BaseModel):
    aspect: str
    document1: str
    document2: str
    recommendation: str


class ComparisonResponse(BaseModel):
    key_differences: list[ComparisonDifference]
    summary: str
    recommendation: str


@router.post("/comparison", response_model=ComparisonResponse)
async def compare_documents(
    request: ComparisonRequest,
    storage: StorageService = Depends(get_storage),
):
    """Compare two SOW documents."""
    doc1 = storage.get_document(request.document_id_1)
    doc2 = storage.get_document(request.document_id_2)

    if not doc1:
        raise HTTPException(status_code=404, detail=f"Document {request.document_id_1} not found")
    if not doc2:
        raise HTTPException(status_code=404, detail=f"Document {request.document_id_2} not found")

    if not doc1.text_content or not doc2.text_content:
        raise HTTPException(status_code=400, detail="Both documents must have extracted text")

    analyzer = OpenAIAnalyzer()
    result = analyzer.get_comparison_analysis(
        text1=doc1.text_content,
        text2=doc2.text_content,
        vendor1=doc1.metadata.vendor_name,
        vendor2=doc2.metadata.vendor_name,
    )

    return ComparisonResponse(
        key_differences=[
            ComparisonDifference(**diff) for diff in result.get("key_differences", [])
        ],
        summary=result.get("summary", ""),
        recommendation=result.get("recommendation", ""),
    )


class ExportRequest(BaseModel):
    analysis_id: str
    include_details: bool = True


@router.post("/export/pdf")
async def export_pdf(
    request: ExportRequest,
    storage: StorageService = Depends(get_storage),
):
    """Export analysis results as PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    analysis = storage.get_analysis(request.analysis_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    document = storage.get_document(analysis.document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
    )

    elements = []

    # Title
    elements.append(Paragraph("SOW Analysis Report", title_style))
    elements.append(Spacer(1, 12))

    # Document info
    elements.append(Paragraph(f"<b>Document:</b> {document.original_filename}", styles['Normal']))
    elements.append(Paragraph(f"<b>Vendor:</b> {document.metadata.vendor_name}", styles['Normal']))
    if document.metadata.contract_value:
        elements.append(Paragraph(f"<b>Contract Value:</b> ${document.metadata.contract_value:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Summary
    if analysis.summary:
        elements.append(Paragraph("<b>Executive Summary</b>", styles['Heading2']))
        elements.append(Paragraph(analysis.summary, styles['Normal']))
        elements.append(Spacer(1, 12))

    # Risk score
    if analysis.risk_score is not None:
        elements.append(Paragraph(f"<b>Overall Risk Score:</b> {analysis.risk_score}/100", styles['Normal']))
    if analysis.total_estimated_savings:
        elements.append(Paragraph(f"<b>Total Estimated Savings:</b> ${analysis.total_estimated_savings:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Findings
    if request.include_details and analysis.findings:
        elements.append(Paragraph("<b>Detailed Findings</b>", styles['Heading2']))
        elements.append(Spacer(1, 10))

        for i, finding in enumerate(analysis.findings, 1):
            elements.append(Paragraph(f"<b>{i}. {finding.title}</b>", styles['Heading3']))
            elements.append(Paragraph(f"Category: {finding.category.value}", styles['Normal']))
            elements.append(Paragraph(f"Risk Level: {finding.risk_level.value.upper()}", styles['Normal']))
            elements.append(Paragraph(f"Description: {finding.description}", styles['Normal']))
            elements.append(Paragraph(f"Evidence: \"{finding.evidence}\"", styles['Normal']))
            elements.append(Paragraph(f"Recommendation: {finding.recommendation}", styles['Normal']))
            if finding.estimated_savings:
                elements.append(Paragraph(f"Estimated Savings: ${finding.estimated_savings:,.2f}", styles['Normal']))
            elements.append(Spacer(1, 15))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=sow_analysis_{analysis.id[:8]}.pdf"
        }
    )


@router.post("/export/excel")
async def export_excel(
    request: ExportRequest,
    storage: StorageService = Depends(get_storage),
):
    """Export analysis results as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    analysis = storage.get_analysis(request.analysis_id)
    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis not found")

    document = storage.get_document(analysis.document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = "SOW Analysis Report"
    ws_summary['A1'].font = Font(size=16, bold=True)

    ws_summary['A3'] = "Document"
    ws_summary['B3'] = document.original_filename
    ws_summary['A4'] = "Vendor"
    ws_summary['B4'] = document.metadata.vendor_name
    ws_summary['A5'] = "Contract Value"
    ws_summary['B5'] = f"${document.metadata.contract_value:,.2f}" if document.metadata.contract_value else "N/A"
    ws_summary['A6'] = "Risk Score"
    ws_summary['B6'] = f"{analysis.risk_score}/100" if analysis.risk_score else "N/A"
    ws_summary['A7'] = "Total Findings"
    ws_summary['B7'] = len(analysis.findings)
    ws_summary['A8'] = "Estimated Savings"
    ws_summary['B8'] = f"${analysis.total_estimated_savings:,.2f}" if analysis.total_estimated_savings else "N/A"

    # Findings sheet
    ws_findings = wb.create_sheet("Findings")

    headers = ["#", "Title", "Category", "Risk Level", "Description", "Evidence", "Recommendation", "Est. Savings"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws_findings.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row, finding in enumerate(analysis.findings, 2):
        ws_findings.cell(row=row, column=1, value=row - 1)
        ws_findings.cell(row=row, column=2, value=finding.title)
        ws_findings.cell(row=row, column=3, value=finding.category.value)
        ws_findings.cell(row=row, column=4, value=finding.risk_level.value)
        ws_findings.cell(row=row, column=5, value=finding.description)
        ws_findings.cell(row=row, column=6, value=finding.evidence)
        ws_findings.cell(row=row, column=7, value=finding.recommendation)
        ws_findings.cell(row=row, column=8, value=finding.estimated_savings or "")

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws_findings.column_dimensions[chr(64 + col)].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=sow_analysis_{analysis.id[:8]}.xlsx"
        }
    )
